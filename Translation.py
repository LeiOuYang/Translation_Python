#!/usr/bin/env python

### excel表格翻譯

import os
import json
import urllib
import time,threading
import xml.etree.ElementTree as ET
from urllib import request,parse
from openpyxl import *

"""-----------------------------------------------------------------"""
### 功能说明
def information():
    print(
    """
-----------------------------------------------------------------
            <<< 表格自动翻译工具 >>>

    使用说明:

        1、将需要翻译的表格文件防止脚本同级目录下，文件格式为 xlsx
        2、windows用户可运行目录下的run.bat文件
        3、输入文件名称，不需要输入文件格式
        4、输入需要翻译的表格单元格列名，例如： A，B
        5、输入翻译保存表格单元格
        6、直到运行完成即可(提示' 翻译完成 '字段信息)

    注意：目前只支持英文自动翻译成中文
    
    作者：Awesome       QQ：2281280195
------------------------------------------------------------------
    """
        )
"""-----------------------------------------------------------------"""

"""-----------------------------------------------------------------"""
### 百度翻译接口，如果翻译不成功，返回字符串 'THRANS_ERROR'
### 翻译成功，返回翻译字段
def BaiDuTranslate(appid, secretKey, fromLang, toLang, src_trans):
    #print('Python test baidu fanyi api')
    
    s_url = 'http://api.fanyi.baidu.com/api/trans/vip/translate'  #提交地址
    salt = random.randint(32768, 65536)  #随机数
    sign = appid+src_trans+str(salt)+secretKey
    
    #计算MD5校验码
    o_md5 = hashlib.md5()
    o_md5.update(sign.encode('utf-8'))
    sign = o_md5.hexdigest()
    #end
    
    #格式化http有效数据
    #q = parse.quote_plus(q)
    data = {
        'appid': appid,
        'q': src_trans,
        'from': fromLang,
        'to': toLang,
        'salt': str(salt),
        'sign': sign
    }
    #end
    
    #编码数据并且POST提交，成功返回json数据，并进行解析
    data = parse.urlencode(data)
    data = data.encode('utf8')
    req = request.Request(s_url, data=data, method='POST')
    res = request.urlopen(req)

    #print(res.read().decode('utf8'))

    if res.status==200:
        json_s = res.read().decode('utf8')
        json_d = json.loads(json_s)
        if 'trans_result' in json_d.keys():
            result = json_d['trans_result']
            dst = result[0]
            dst = dst['dst']
            return str(dst)
        else:
            return 'THRANS_ERROR'
### 百度翻译接口函数定义结束
"""-----------------------------------------------------------------"""
            
"""-------------------------------------------------------------------"""
### 输入待处理的表格文件名称，不带格式
### 代码功能：
###     1、实现表格复制，并新保存表格
###     2、表格不存在，会输出提醒用户信息
def create_xlsx(src_filename):
    temp_filename = src_filename +'.xlsx'
    if not(os.path.isfile(temp_filename) and os.path.exists(temp_filename)):
        print('source file no exists...')
        return
    dest_filename = src_filename+'_temp.xlsx'
    src_wb = load_workbook(filename=temp_filename)
    dest_wb = Workbook()

    #复制整个表格至新表中
    for sheet in src_wb:
        print(sheet.title)
        dest_wb.create_sheet(sheet.title)
        src_ws = src_wb[sheet.title]
        dest_ws = dest_wb[sheet.title]
        row_len = len(list(src_ws.rows))
        col_len = len(list(src_ws.columns))
        
        for row in range(1,row_len+1):
            for col in range(1,col_len+1):
                cell = src_ws.cell(row=row, column=col) #获取表格中单元格
                dest_ws.cell(row=row, column=col, value=cell.value)
                col += 1
            row += 1
    del dest_wb['Sheet']
    dest_wb.save(dest_filename)
    return (row_len, col_len)
"""-------------------------------------------------------------------"""

"""-------------------------------------------------------------------"""
### 各平台翻译函数接口
### 输入参数：输入两个元素的元组数据，分别为：翻译平台地址，需要翻译的数据(初始语言，目标语言，待翻译值)
### 返回数据：返回一个字典数据{翻译平台：翻译结果}
"""
#谷歌翻译接口返回一个txt文件
    #http://translate.google.cn/translate_a/single?client=gtx&dt=t&dj=1&ie=UTF-8&sl=auto&tl=zh_TW&q=calculate
#微软翻译接口返回的是xml文档
    #http://api.microsofttranslator.com/v2/Http.svc/Translate?appId=AFC76A66CF4F434ED080D245C30CF1E71C22959C&from=&to=en&text=考勤计算
#有道翻译接口返回json格式数据
    #http://fanyi.youdao.com/translate?&doctype=json&type=AUTO&i=计算
#百度翻译接口返回json格式数据
    #http://fanyi.baidu.com/transapi?from=auto&to=cht&query=Calculation
"""
def translate_data(tdata=('BAIDU', ('en', 'zh', ''))):

    if tdata==() or len(tdata)!=2:
        return {}

    api_class = tdata[0].strip().upper()
    api_url = ''
    api_data = {}
    dest_data = ''
    if 'BAIDU'==api_class:
        api_url = 'http://fanyi.baidu.com/transapi?'
        value = list(tdata[1])
        api_data = { 'from':'auto', 'to':value[1], 'query':value[2] }
    elif 'YOUDAO'==api_class:
        api_url = 'http://fanyi.youdao.com/translate?'
        value = list(tdata[1])
        api_data = { 'doctype':'json', 'type':'auto', 'i':value[2] }
    elif 'BING'==api_class:
        api_url = 'http://api.microsofttranslator.com/v2/Http.svc/Translate?appId=AFC76A66CF4F434ED080D245C30CF1E71C22959C'
        value = list(tdata[1])
        api_data = { 'from':'', 'to':value[1], 'text':value[2] }
    elif 'GOOGLE'==api_class:
        api_url = 'http://translate.google.cn/translate_a/single?client=gtx'
        value = list(tdata[1])
        api_data = { 'dt':'t',  'dj':'1', 'ie':'UTF-8', 'sl':value[0], 'tl':value[1], 'q':value[2] }
    else:
        return {}
    
    #定义http头部
    headers = {
        'Accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        #'Accept-Encoding':'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
        'Connection':'keep-alive',
        #'Host' : 'fanyi.baidu.com',
        'Upgrade-Insecure-Requests':'1',
        'User-Agent' : 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:64.0) Gecko/20100101 Firefox/64.0'
    }

    #获取http翻译应答数据
    data = parse.urlencode(api_data)
    data = data.encode('utf8')
    data_str = '&'+data.decode('utf8')
    req = request.Request(url=api_url+data_str, headers=headers, method="GET")
    res_data = request.urlopen(req)
    if res_data.status==200:
        res_data = res_data.read().decode('utf8')
    else:
        return {}

    #解析获取翻译字符串
    if 'BAIDU'==api_class:  #json处理
        json_d = json.loads(res_data)
        if 'data' in json_d.keys():
            data = json_d['data']
            data_dic = data[0]
            dest_data = data_dic['dst']
        else:
            return {}
    elif 'YOUDAO'==api_class: #json处理
        json_d = json.loads(res_data)
        if 'translateResult' in json_d.keys():
            data = json_d['translateResult']
            data= data[0]
            data_dic = data[0]
            dest_data = data_dic['tgt']
        else:
            return {}
    elif 'GOOGLE'==api_class:
        json_d = json.loads(res_data)
        if 'sentences' in json_d.keys():
            data = json_d['sentences']
            data_dic = data[0]
            dest_data = data_dic['trans']
        else:
            return {}
    elif 'BING'==api_class:
        root = ET.fromstring(res_data)
        dest_data = root.text
        
    return {api_class:dest_data}
        

"""-------------------------------------------------------------------"""

"""-------------------------------------------------------------------"""
### 线程函数执行接口
### src=待翻译列 dest=翻译保存列  start_count-开始行数  end_count-结束行数
def baidu_thread_loop(src, dest, start_count, end_count):
    while True:
        time.sleep(0.01)
        

"""-------------------------------------------------------------------"""


### 主程序代码
def user_main():
    
    #全局数据
    api_class = ''   #翻译接口   baidu，google，bing，youdao
    wb_max_rows = 0  #最大行 
    wb_max_col = 0   #最大列
    wb_src_filename = '' #待表格翻译文件名
    enCell = None
    zhCell = None
    API_CLASS = ('BAIDU','YOUDAO','GOOGLE', 'BING')

    baidu_start_pos = 1

    information()    #显示提示信息

    while True:
        api_class = input('选择使用的翻译平台：(baidu、google、youdao、bing) ')
        api_class = api_class.strip().upper()
        if api_class in API_CLASS:
            print('\t翻译平台 '+api_class.lower())
            break

    while True:
        filename = input("请输入表格的名称:  ")
        if not(os.path.isfile(filename+'.xlsx') and os.path.exists(filename+'.xlsx')):
            continue
        print("\t"+filename+'.xlsx'+"存在")
        wb_src_filename = filename+'.xlsx'
        t_len = create_xlsx(filename)  #创建表格文件，并复制该表格
        wb_max_rows = t_len[0]
        wb_max_col = t_len[1]
        baidu_count = 10000
        if baidu_count<=wb_max_rows:
            baidu_count = wb_max_rows
        break
    while True:    
        fromCell = input("待翻译列：(如：A列)  ")
        if not fromCell.isalpha():
            continue
        fromCell = fromCell.strip().upper()
        print("\t"+"输入>>" + fromCell)
        break
    while True:
        toCell = input("翻译保存列:  ")
        toCell = toCell.strip().upper()
        if not toCell.isalpha():
            continue
        if toCell==fromCell:
            continue
        print("\t"+"输入>>" + toCell)
        break
    print("输入数据完成")

    wb_src_filename = filename+'_temp.xlsx'
    wb_src = load_workbook(wb_src_filename)  #加载工作表
    ws_sheet = wb_src[wb_src.sheetnames[0]]  #取工作表中第一张表
    print('rows:'+str(wb_max_rows) + '  col: '+str(wb_max_col) + ' sheet: '+ws_sheet.title)

    row = baidu_start_pos   #开始翻译行开始数
    for row in range(baidu_start_pos,baidu_count+1):
        error_count = 0
        result = ''
        fromCell_t = str(fromCell)+str(row)
        toCell_t = str(toCell)+str(row)
        cell = ws_sheet[fromCell_t]
        srcs = cell.value
        cell = ws_sheet[toCell_t]
        dests = cell.value
        if (srcs is None) or (''==srcs.strip()) or ((dests!=None) and (''!=dests.strip())): 
            row += 1
            continue
        while True:
            result = translate_data(tdata=(api_class, ('en', 'zh', srcs)))
            print(result)
            if {}==result:
                error_count += 1
                continue
                if error_count>10:
                    print(fromCell_t + ' translate error...')
                    break
            else:
                break
        cell.value = ''
        cell.value += result[api_class]
        row += 1
        time.sleep(0.01)
    wb_src.save(wb_src_filename)
    print("\n\t---<<< 翻译工作完成 >>>---")
    
    """
        
        
    #多线程处理
    
    #baidu_thread = threading.Thread(target=baidu_thread_loop, name="BAIDU_API", args=('baidu',12 ))
    #baidu_thread.start()
   """
while True:
    user_main()
    print('\n\n')
    print('-'*60)

